## LIBRARIES:
import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog
from tkinter.filedialog import askdirectory
import numpy as np
from sentence_transformers import SentenceTransformer, util
import torch
import unicodedata
import re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sentence_transformers import CrossEncoder
import shutil
import re
MODEL_PATH = r"C:\Users\mshanmugam\.cache\huggingface\hub\models--sentence-transformers--all-MiniLM-L6-v2\snapshots\1110a243fdf4706b3f48f1d95db1a4f5529b4d41"

CROSS_PATH = r"C:\Users\mshanmugam\.cache\huggingface\hub\models--cross-encoder--stsb-distilroberta-base\snapshots\6b71347df6e2b34246b53e06d6bce70ef67de368"

class initial_class:
    def __init__(self, bi_encoder_name="all-MiniLM-L6-v2", cross_encoder_name="cross-encoder/stsb-distilroberta-base"):
        self.model = SentenceTransformer(MODEL_PATH)
        self.cross_model = CrossEncoder(CROSS_PATH)
    # ============================================================
    # ✅ CONTENT TYPE INPUT
    # ============================================================

    def ask_content_type(self):
        while True:
            value = input("Enter template type [series/movie]: ").strip().lower()
            if value in {"series", "movie"}:
                return value
            print("❌ Invalid input. Please enter either 'series' or 'movie'.")


    # ============================================================
    # ✅ GENERIC HELPERS
    # ============================================================
    @staticmethod
    def first_existing_column(df, candidates, required=True):
        """
        Return the first matching column from candidates.
        """
        for col in candidates:
            if col in df.columns:
                return col
        if required:
            raise ValueError(f"None of these columns found: {candidates}")
        return None


    @staticmethod
    def get_output_title_column(df):
        """
        For final output/template creation, find the best title column.
        """
        candidates = [
            "*Title name",
            "wb2b epiosde name",
            "Movie Title",
            "Title",
            "main-title"
        ]
        for c in candidates:
            if c in df.columns:
                return c
        return None


    @staticmethod
    def pick_input_folder(title="Select folder with WBTV / WB2B / Foundry files"):
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)   # brings dialog to front (Windows helpful)
        root.update()                      # IMPORTANT: forces Tk to initialize properly
        folder = filedialog.askdirectory(title=title)
        root.destroy()
        return folder
    
    
    @staticmethod
    def select_output_folder():
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        root.update()

        print("Select folder to save output files...")
        folder_path = filedialog.askdirectory(title="Select folder to save output files")
        root.destroy()

        if not folder_path:
            print("No folder selected!")
            return None

        # ✅ Clear folder contents
        for item in os.listdir(folder_path):
            item_path = os.path.join(folder_path, item)

            try:
                if os.path.isfile(item_path) or os.path.islink(item_path):
                    os.remove(item_path)   # delete file
                elif os.path.isdir(item_path):
                    shutil.rmtree(item_path)  # delete folder
            except Exception as e:
                print(f"❌ Failed to delete {item_path}: {e}")

        print("✅ Folder cleaned successfully")
        print("Selected Output Folder:", folder_path)

        return folder_path

    # ============================================================
    # 🔒 TEXT NORMALIZATION (UNCHANGED)
    # ============================================================

    @staticmethod
    def normalize_title(title):
        if pd.isna(title):
            return ""
        s = str(title)

        # normalize unicode
        s = unicodedata.normalize("NFKC", s)

        # handle Excel hidden spaces
        s = s.replace("\u00A0", " ")        # NBSP
        s = s.replace("\u200B", "")         # zero-width space
        s = s.replace("\uFEFF", "")         # BOM

        s = s.strip().lower()
        s = re.sub(r"[^\w\s]", "", s)       # remove punctuation
        s = re.sub(r"\b\d{4}\b", "", s)     # remove years
        s = re.sub(r"\s+", " ", s)          # collapse spaces
        return s


    # ============================================================
    # ✅ SOURCE DETECTION
    # ============================================================
    @staticmethod
    def detect_sources(input_folder: str):
        files = os.listdir(input_folder)
        sources = {}
        for f in files:
            lf = f.lower()
            full = os.path.join(input_folder, f)
            if "wbtv" in lf:
                sources["WBTV"] = full
            elif "wb2b" in lf:
                sources["WB2B"] = full
            elif "foundry" in lf:
                sources["FOUNDRY"] = full
        return sources

class load_and_match(initial_class):
    # ============================================================
    # ✅ LOAD FOUNDRY (SERIES + MOVIE SUPPORT)
    # ============================================================
    def load_foundry(self, foundry_path: str, content_type: str = "series"):
        
        ext = os.path.splitext(foundry_path)[-1].lower()

        if ext == ".csv":
            foundry = pd.read_csv(foundry_path)

        elif ext in [".xlsx", ".xls"]:
            foundry = pd.read_excel(foundry_path)

        else:
            raise ValueError("Unsupported file type! Use CSV or Excel.")


        # Make season / episode safe if columns exist
        if "season-number" in foundry.columns:
            foundry["Season"] = pd.to_numeric(foundry["season-number"], errors="coerce").fillna(0).astype(int)
        else:
            foundry["Season"] = 0

        if "episode-number" in foundry.columns:
            foundry["Episode"] = pd.to_numeric(foundry["episode-number"], errors="coerce").fillna(0).astype(int)
        else:
            foundry["Episode"] = 0
        if "main-title" not in foundry.columns:
            raise KeyError(f"'main-title' not found in Foundry columns: {foundry.columns.tolist()}")

        foundry["*Title name"] = foundry["main-title"]
        foundry["MPM Number"] = ""
        foundry["Primary Release Date"] = pd.NaT

        base_cols = [
            "uuid",
            "asset-type",
            "main-title",
            "*Title name",
            "MPM Number",
            "Primary Release Date",
            "description-400",
            "listings-description",
            "long-description",
            "main-description",
            "short-description",
            "Season",
            "Episode"
        ]

        existing_cols = [c for c in base_cols if c in foundry.columns]
        foundry = foundry[existing_cols].copy()
        asset_type_series = {"series", "season", "episode"}
        asset_type_movie = {"movie", "feature", "film"}

        if content_type == "series":
            series_season = foundry[foundry["asset-type"].isin(["series", "season"])].copy()
            episodes = foundry[foundry["asset-type"] == "episode"].copy()
            movies = pd.DataFrame()
        else:
            movies = foundry[foundry["asset-type"].str.lower().isin(asset_type_movie)].copy()
            if movies.empty:
                # fallback: if asset-type is inconsistent, use rows with Episode == 0
                movies = foundry[foundry["Episode"] == 0].copy()
            series_season = pd.DataFrame()
            episodes = pd.DataFrame()
        

        return foundry, series_season, episodes, movies


    # ============================================================
    # 🔒 WBTV LOADER (SERIES + MOVIE SUPPORT)
    # ============================================================
    def load_wbtv(self, input_wbtv: str, content_type: str = "series", series_season: pd.DataFrame = None):
        ext = os.path.splitext(input_wbtv)[-1].lower()

        if ext == ".csv":
            wbtv = pd.read_csv(input_wbtv, parse_dates=True)

        elif ext in [".xlsx", ".xls"]:
            wbtv = pd.read_excel(input_wbtv)

        else:
            raise ValueError("Unsupported file type! Use CSV or Excel.")

        if content_type == "series":
            columns = [
                "Parent Season Number", "Season/Episode Number",
                "*Title name", "Primary Release Date",
                "MPM Number", "Synopsis (log line)",
                "Synopsis (180)", "Synopsis (full)", "Synopsis (60)"
            ]
            existing_cols = [c for c in columns if c in wbtv.columns]
            wbtv = wbtv[existing_cols].copy()

            if "Parent Season Number" not in wbtv.columns:
                wbtv["Parent Season Number"] = 0
            if "Season/Episode Number" not in wbtv.columns:
                wbtv["Season/Episode Number"] = 0

            mask = wbtv["Parent Season Number"] == 0
            wbtv.loc[mask, ["Parent Season Number", "Season/Episode Number"]] = \
                wbtv.loc[mask, ["Season/Episode Number", "Parent Season Number"]].values

            wbtv[["Parent Season Number", "Season/Episode Number"]] = \
                wbtv[["Parent Season Number", "Season/Episode Number"]].fillna(0)

            wbtv = wbtv.sort_values(by=["Parent Season Number", "Season/Episode Number"])

            wbtv = wbtv.rename(columns={
                "Parent Season Number": "Season",
                "Season/Episode Number": "Episode"
            })

            if series_season is not None and not series_season.empty:
                return pd.merge(
                    wbtv,
                    series_season,
                    on=["Season", "Episode"],
                    how="left",
                    suffixes=("", "_foundry")
                )

            return wbtv

        else:
            
            # MOVIE MODE
            title_col = self.first_existing_column(
                wbtv,
                ["*Title name", "Movie Title", "Title", "Program Title"],
                required=True
            )

            date_col = self.first_existing_column(
                wbtv,
                ["Primary Release Date", "Release Date", "Original Release Date"],
                required=False
            )

            mpm_col = self.first_existing_column(
                wbtv,
                ["MPM Number", "Title ID", "Internal Reference"],
                required=False
            )

            synopsis_candidates = [
                "Synopsis (log line)", "Synopsis (180)", "Synopsis (full)", "Synopsis (60)"
            ]
            keep_cols = [title_col]
            if date_col:
                keep_cols.append(date_col)
            if mpm_col:
                keep_cols.append(mpm_col)
            keep_cols.extend([c for c in synopsis_candidates if c in wbtv.columns])

            wbtv = wbtv[keep_cols].copy()
            wbtv["Season"] = 0
            wbtv["Episode"] = 0

            rename_map = {title_col: "*Title name"}
            if date_col and date_col != "Primary Release Date":
                rename_map[date_col] = "Primary Release Date"
            if mpm_col and mpm_col != "MPM Number":
                rename_map[mpm_col] = "MPM Number"

            wbtv = wbtv.rename(columns=rename_map)
            return wbtv


    # ============================================================
    # 🔒 WB2B LOADER (SERIES + MOVIE SUPPORT)
    # ============================================================
    def load_wb2b(self,input_wb2b: str, content_type: str = "series", series_season: pd.DataFrame = None, has_wbtv: bool = False):
        ext = os.path.splitext(input_wb2b)[-1].lower()

        if ext == ".csv":
            wb2b = pd.read_csv(input_wb2b, parse_dates=True)

        elif ext in [".xlsx", ".xls"]:
            wb2b = pd.read_excel(input_wb2b)

        else:
            raise ValueError("Unsupported file type! Use CSV or Excel.")

        if content_type == "series":
            if "Series MPM #" in wb2b.columns and "Season Year" in wb2b.columns:
                wb2b.loc[wb2b["Series MPM #"].isna(), "Season Year"] = 0

            columns = [
                "Season Year", "Episode Number", "Episode Name",
                "Original Broadcast Date", "Title ID",
                "Short Synopsis", "Synopsis"
            ]
            existing_cols = [c for c in columns if c in wb2b.columns]
            wb2b = wb2b[existing_cols].copy()

            if "Episode Number" not in wb2b.columns:
                wb2b["Episode Number"] = 0
            if "Season Year" not in wb2b.columns:
                wb2b["Season Year"] = 0

            wb2b["Episode Number"] = wb2b["Episode Number"].fillna(0)

            wb2b = wb2b.rename(columns={
                "Short Synopsis": "Synopsis (60)",
                "Synopsis": "Synopsis (full)",
                "Season Year": "Season",
                "Episode Number": "Episode",
                "Episode Name": "wb2b epiosde name",
                "Original Broadcast Date": "Primary Release Date",
                "Title ID": "MPM Number"
            })

            if "Primary Release Date" in wb2b.columns:
                wb2b["Primary Release Date"] = pd.to_datetime(
                    wb2b["Primary Release Date"], errors="coerce"
                ).dt.strftime("%Y-%m-%d")

            wb2b = wb2b.sort_values(by=["Season", "Episode"])

            if (not has_wbtv) and (series_season is not None) and (not series_season.empty):
                return pd.merge(
                    wb2b,
                    series_season,
                    on=["Season", "Episode"],
                    how="left",
                    suffixes=("", "_foundry"))
            return wb2b

        else:
            # MOVIE MODE
            title_col = self.first_existing_column(
                wb2b,
                ["Episode Name", "Movie Title", "Title", "Program Title"],
                required=True
            )
            date_col = self.first_existing_column(
                wb2b,
                ["Original Broadcast Date", "Release Date", "Primary Release Date"],
                required=False
            )
            id_col = self.first_existing_column(
                wb2b,
                ["Title ID", "MPM Number", "Internal Reference"],
                required=False
            )

            keep_cols = [title_col]
            if date_col:
                keep_cols.append(date_col)
            if id_col:
                keep_cols.append(id_col)
            keep_cols.extend([c for c in ["Short Synopsis", "Synopsis"] if c in wb2b.columns])

            wb2b = wb2b[keep_cols].copy()

            wb2b = wb2b.rename(columns={
                title_col: "wb2b epiosde name",
                "Short Synopsis": "Synopsis (60)",
                "Synopsis": "Synopsis (full)",
                date_col if date_col else "": "Primary Release Date",
                id_col if id_col else "": "MPM Number"
            })

            wb2b["Season"] = 0
            wb2b["Episode"] = 0

            if "Primary Release Date" in wb2b.columns:
                wb2b["Primary Release Date"] = pd.to_datetime(
                    wb2b["Primary Release Date"], errors="coerce"
                ).dt.strftime("%Y-%m-%d")

            return wb2b


    # ============================================================
    # 🔧 MATCHING HELPERS
    # ============================================================
    @staticmethod
    def extract_number(text):
        match = re.search(r"\d+", str(text))
        return int(match.group()) if match else None

    @staticmethod
    def is_generic_episode(title):
        if pd.isna(title):
            return False

        t = str(title).strip()

        pattern = r"""
        ^\s*(
            (episode|ep)\.?\s*[-:#]?\s*\d+      # Episode 1 / Ep. 1 / Episode #1
            |
            e\s*[-:#]?\s*\d+                   # E01 / E-01 / E:01
            |
            s\s*\d+\s*e\s*\d+                  # S1E1 / S01E01 / S1 E1
        )\s*([:.\-–—].*)?$                     # allow trailing ": Pilot" or "- Pilot"
        """
        return bool(re.match(pattern, t, flags=re.IGNORECASE | re.VERBOSE))


    # ============================================================
    # ✅ MOVIE TITLE-ONLY MATCHER
    # ============================================================
    def semantic_movie_match_and_merge(self,
        left_df: pd.DataFrame,
        right_df: pd.DataFrame,
        *,
        normalize_title_fn,
        left_id_col: str,
        left_title_col: str,
        right_title_col: str,
        right_pull_cols=(
            "uuid", "asset-type", "main-title",
            "description-400", "listings-description",
            "long-description", "main-description",
            "short-description", "Season", "Episode"
        ),
        top_k: int = 5,
        ce_threshold: float = 0.75,
        model=None,
        cross_model=None
    ):
        if model is None:
            model = self.model
        if cross_model is None:
            cross_model = self.cross_model

        L = left_df.copy().reset_index(drop=True)
        R = right_df.copy().reset_index(drop=True)

        L["norm_title"] = L[left_title_col].apply(normalize_title_fn)
        R["norm_title"] = R[right_title_col].apply(normalize_title_fn)

        L_emb = model.encode(L["norm_title"].tolist(), convert_to_tensor=True)
        R_emb = model.encode(R["norm_title"].tolist(), convert_to_tensor=True)
        cos_scores = util.cos_sim(L_emb, R_emb)

        right_pull_cols = list(right_pull_cols)

        matches = {left_id_col: []}
        for c in right_pull_cols:
            matches[c] = []
        matches["match_type"] = []

        used_right = set()

        for i in L.index:
            scores = cos_scores[i]
            top_pos = torch.topk(scores, k=min(top_k, len(scores))).indices.tolist()
            cand_idx = [j for j in top_pos if j not in used_right]

            if not cand_idx:
                continue

            pairs = [(L.loc[i, "norm_title"], R.loc[j, "norm_title"]) for j in cand_idx]
            ce_scores_arr = cross_model.predict(pairs)

            best_pos = int(np.argmax(ce_scores_arr))
            best_ce = ce_scores_arr[best_pos]
            best_j = cand_idx[best_pos]

            if best_ce >= ce_threshold:
                matches[left_id_col].append(L.loc[i, left_id_col])
                for c in right_pull_cols:
                    matches[c].append(R.loc[best_j, c] if c in R.columns else np.nan)
                matches["match_type"].append("TITLE ONLY Semantic")
                used_right.add(best_j)

        matches_df = pd.DataFrame(matches)
        merged = L.merge(matches_df, on=left_id_col, how="left", suffixes=("", "_df2"))

        for col in matches_df.columns:
            if col == left_id_col:
                continue
            df2_col = f"{col}_df2"
            if df2_col in merged.columns:
                merged[col] = merged[col].combine_first(merged[df2_col])

        merged = merged[[c for c in merged.columns if not c.endswith("_df2")]]
        return merged, matches_df, used_right, model, cross_model

class matching_pipeline(load_and_match):
    # ============================================================
    # ✅ MOVIE WRAPPERS
    # ============================================================
    def match_wbtv_to_foundry_movies(self, WBTV_0, Movies, model=None, cross_model=None, top_k=5, ce_threshold=0.75):
        return self.semantic_movie_match_and_merge(
            left_df=WBTV_0,
            right_df=Movies,
            normalize_title_fn=self.normalize_title,
            left_id_col="MPM Number",
            left_title_col="*Title name",
            right_title_col="main-title",
            right_pull_cols=(
                "uuid", "asset-type", "main-title",
                "description-400", "listings-description",
                "long-description", "main-description", "short-description",
                "Season", "Episode"
            ),
            top_k=top_k,
            ce_threshold=ce_threshold,
            model=model,
            cross_model=cross_model,
        )


    def match_wb2b_to_foundry_movies(self, WB2B_0, Movies, model=None, cross_model=None, top_k=5, ce_threshold=0.75):
        return self.semantic_movie_match_and_merge(
            left_df=WB2B_0,
            right_df=Movies,
            normalize_title_fn=self.normalize_title,
            left_id_col="MPM Number",
            left_title_col="wb2b epiosde name",
            right_title_col="main-title",
            right_pull_cols=(
                "uuid", "asset-type", "main-title",
                "description-400", "listings-description",
                "long-description", "main-description", "short-description",
                "Season", "Episode"
            ),
            top_k=top_k,
            ce_threshold=ce_threshold,
            model=model,
            cross_model=cross_model,
        )


    def match_wbtv_with_wb2b_movies(self,
        WBTV_0: pd.DataFrame,
        WB2B_0: pd.DataFrame,
        model=None,
        cross_model=None,
        top_k: int = 5,
        ce_threshold: float = 0.75,
    ):
        wb2b_pull_cols = [
            "wb2b epiosde name",
            "Synopsis (60)",
            "Synopsis (full)",
            "Season",
            "Episode"
        ]

        WBTV_0 = WBTV_0.copy()
        WBTV_0["_wbtv_order"] = np.arange(len(WBTV_0))

        final, matches_df, used_right, model, cross_model = self.semantic_movie_match_and_merge(
            left_df=WBTV_0,
            right_df=WB2B_0,
            normalize_title_fn=self.normalize_title,
            left_id_col="MPM Number",
            left_title_col="*Title name",
            right_title_col="wb2b epiosde name",
            right_pull_cols=wb2b_pull_cols,
            top_k=top_k,
            ce_threshold=ce_threshold,
            model=model,
            cross_model=cross_model,
        )

        final = final.sort_values("_wbtv_order").drop(columns="_wbtv_order").reset_index(drop=True)
        return final, matches_df, used_right, model, cross_model


    # ============================================================
    # ✅ GENERIC SEMANTIC + CE MATCHER (REUSABLE)
    # ============================================================
    def semantic_episode_match_and_merge(self,
        left_df: pd.DataFrame,
        right_df: pd.DataFrame,
        *,
        normalize_title_fn,
        left_id_col: str,
        left_title_col: str,
        right_title_col: str,
        season_col: str = "Season",
        episode_col: str = "Episode",
        right_pull_cols=(
            "uuid", "asset-type", "main-title",
            "description-400", "listings-description",
            "long-description", "main-description",
            "short-description", "Season", "Episode"
        ),
        top_k: int = 5,
        ce_threshold: float = 0.60,
        fallback_cosine_threshold: float = 0.30,   # FIX 3: quality gate on fallback
        exclude_left_episode_zero: bool = True,
        model=None,
        cross_model=None
    ):
        """
        Matches left_df rows to right_df rows using:
        1) SAME SEASON  — cosine topK + cross-encoder (stsb 0-1 scores)
        2) ALL SEASONS  — cosine topK + cross-encoder
        3) FALLBACK     — exact Season+Episode with minimum cosine gate
        Merges right_pull_cols + match_type back into left_df using left_id_col.
        """

        # ── init models once ──────────────────────────────────────────────────────
        if model is None:
            model = self.model
        if cross_model is None:
            cross_model = self.cross_model

        L = left_df.copy()
        R = right_df.copy()

        # ── exclude series/season rows from left matching pool ────────────────────
        if exclude_left_episode_zero and episode_col in L.columns:
            L_match = L[L[episode_col] != 0].copy()
        else:
            L_match = L.copy()

        L_match = L_match.reset_index(drop=True)
        R = R.reset_index(drop=True)

        # ── normalize titles ──────────────────────────────────────────────────────
        L_match["norm_title"] = L_match[left_title_col].apply(normalize_title_fn)
        R["norm_title"]       = R[right_title_col].apply(normalize_title_fn)

        # ── numeric + generic flags ───────────────────────────────────────────────
        L_match["num"]        = L_match[left_title_col].apply(self.extract_number)
        R["num"]              = R[right_title_col].apply(self.extract_number)
        L_match["is_generic"] = L_match[left_title_col].apply(self.is_generic_episode)
        R["is_generic"]       = R[right_title_col].apply(self.is_generic_episode)

        # ── encode all titles once ────────────────────────────────────────────────
        L_emb     = model.encode(L_match["norm_title"].tolist(), convert_to_tensor=True)
        R_emb     = model.encode(R["norm_title"].tolist(),       convert_to_tensor=True)
        cos_scores = util.cos_sim(L_emb, R_emb)          # shape: [len(L_match), len(R)]

        # ── FIX 4: rename Season/Episode in right_pull_cols before collecting ─────
        # This prevents combine_first from overwriting L's Season/Episode with R's.
        right_pull_cols = list(right_pull_cols)
        renamed_pull_cols = [
            "foundry_Season" if c == season_col else
            "foundry_Episode" if c == episode_col else c
            for c in right_pull_cols
        ]

        # ── match store ───────────────────────────────────────────────────────────
        matches = {left_id_col: []}
        for c in renamed_pull_cols:
            matches[c] = []
        matches["match_type"] = []

        used_right = set()

        # ── main matching loop ────────────────────────────────────────────────────
        for i in L_match.index:
            season  = int(L_match.loc[i, season_col])
            episode = int(L_match.loc[i, episode_col])
            wtitle  = L_match.loc[i, "norm_title"]
            found   = False
            best_j  = None
            match_type = None

            # FIX 2: only apply numeric conflict check for generic episode titles
            # e.g. "Episode 3" vs "Episode 5" should block; "24 Hours…" vs "24 Hours…" should not
            num1 = L_match.loc[i, "num"] if L_match.loc[i, "is_generic"] else None

            def num_conflict(j):
                """True if both sides have numbers and they differ."""
                num2 = R.loc[j, "num"] if R.loc[j, "is_generic"] else None
                return (num1 is not None and num2 is not None and num1 != num2)

            # ── STEP 1: same season ───────────────────────────────────────────────
            same_season_idx = R[R[season_col] == season].index
            if len(same_season_idx) > 0:
                e_idx_list = list(same_season_idx)                    # FIX: plain list avoids pandas positional ambiguity
                scores     = cos_scores[i][e_idx_list]
                top_pos    = torch.topk(scores, k=min(top_k, len(scores))).indices.tolist()
                cand_idx   = [e_idx_list[p] for p in top_pos if e_idx_list[p] not in used_right]

                if cand_idx:
                    pairs    = [(wtitle, R.loc[j, "norm_title"]) for j in cand_idx]
                    ce_scores_arr = cross_model.predict(pairs)
                    best_pos = int(np.argmax(ce_scores_arr))
                    best_ce  = ce_scores_arr[best_pos]
                    candidate_j = cand_idx[best_pos]

                    if best_ce >= ce_threshold and not num_conflict(candidate_j):
                        found      = True
                        best_j     = candidate_j
                        match_type = "SAME SEASON Semantic"

            # ── STEP 2: all seasons ───────────────────────────────────────────────
            if not found:
                e_idx_list = list(R.index)
                scores     = cos_scores[i][e_idx_list]
                top_pos    = torch.topk(scores, k=min(top_k, len(scores))).indices.tolist()
                cand_idx   = [e_idx_list[p] for p in top_pos if e_idx_list[p] not in used_right]

                if cand_idx:
                    pairs    = [(wtitle, R.loc[j, "norm_title"]) for j in cand_idx]
                    ce_scores_arr = cross_model.predict(pairs)
                    best_pos = int(np.argmax(ce_scores_arr))
                    best_ce  = ce_scores_arr[best_pos]
                    candidate_j = cand_idx[best_pos]

                    if best_ce >= ce_threshold and not num_conflict(candidate_j):
                        found      = True
                        best_j     = candidate_j
                        match_type = "ALL SEASONS Semantic"

            # ── STEP 3: exact Season + Episode fallback ───────────────────────────
            # FIX 3: add minimum cosine gate — fallback no longer accepts what
            # the semantic threshold explicitly rejected
            if not found and episode != 0:
                exact  = R[(R[season_col] == season) & (R[episode_col] == episode)]
                unused = [idx for idx in exact.index if idx not in used_right]

                if unused:
                    scores   = cos_scores[i][unused]
                    best_pos = torch.argmax(scores).item()
                    best_cos = scores[best_pos].item()

                    if best_cos >= fallback_cosine_threshold:          # FIX 3
                        best_j     = unused[best_pos]
                        found      = True
                        match_type = "Season + Episode"

            # ── store result ──────────────────────────────────────────────────────
            if found:
                matches[left_id_col].append(L_match.loc[i, left_id_col])
                for orig_col, renamed_col in zip(right_pull_cols, renamed_pull_cols):
                    matches[renamed_col].append(R.loc[best_j, orig_col])
                matches["match_type"].append(match_type)
                used_right.add(best_j)

        # ── build matches dataframe ───────────────────────────────────────────────
        matches_df = pd.DataFrame(matches)

        merged = L.merge(matches_df, on=left_id_col, how="left", suffixes=("", "_df2"))

        # combine_first logic (same as your merge style)
        for col in matches_df.columns:
            if col == left_id_col:
                continue
            df2_col = f"{col}_df2"
            if df2_col in merged.columns:
                merged[col] = merged[col].combine_first(merged[df2_col])

        merged = merged[[c for c in merged.columns if not c.endswith("_df2")]]

        return merged, matches_df, used_right, model, cross_model


    # ============================================================
    # ✅ SPECIFIC WRAPPERS (JUST TO KEEP NAMES YOU WANT)
    # ============================================================
    def match_wbtv_to_foundry_episodes(self,WBTV_0, Episodes, model=None, cross_model=None, top_k=5, ce_threshold=0.75):
        return self.semantic_episode_match_and_merge(
            left_df=WBTV_0,
            right_df=Episodes,
            normalize_title_fn=self.normalize_title,
            left_id_col="MPM Number",
            left_title_col="*Title name",
            right_title_col="main-title",
            right_pull_cols=(
                "uuid","asset-type","main-title","description-400","listings-description",
                "long-description","main-description","short-description","Season","Episode"
            ),
            top_k=top_k,
            ce_threshold=ce_threshold,
            exclude_left_episode_zero=True,
            model=model,
            cross_model=cross_model,
        )


    def match_wb2b_to_foundry_episodes(self, WB2B_0, Episodes, model=None, cross_model=None, top_k=5, ce_threshold=0.75):
        return self.semantic_episode_match_and_merge(
            left_df=WB2B_0,
            right_df=Episodes,
            normalize_title_fn=self.normalize_title,
            left_id_col="MPM Number",
            left_title_col="wb2b epiosde name",
            right_title_col="main-title",
            right_pull_cols=(
                "uuid","asset-type","main-title","description-400","listings-description",
                "long-description","main-description","short-description","Season","Episode"
            ),
            top_k=top_k,
            ce_threshold=ce_threshold,
            exclude_left_episode_zero=True,
            model=model,
            cross_model=cross_model,
        )


    def match_wbtv_with_wb2b_episodes(self,
        WBTV_0: pd.DataFrame,
        WB2B_0: pd.DataFrame,
        model=None,
        cross_model=None,
        top_k: int = 5,
        ce_threshold: float = 0.75,
    ):
        wb2b_pull_cols = [
            "Season",
            "Episode",
            "wb2b epiosde name",
            "Synopsis (60)",
            "Synopsis (full)",
        ]

        # ✅ Preserve original WBTV order
        WBTV_0 = WBTV_0.copy()
        WBTV_0["_wbtv_order"] = np.arange(len(WBTV_0))

        # -------------------------------------------------
        # 1️⃣ SERIES + SEASON (Episode == 0)
        # -------------------------------------------------
        wbtv_ss = WBTV_0[WBTV_0["Episode"] == 0].copy()
        wbtv_ep = WBTV_0[WBTV_0["Episode"] > 0].copy()

        wb2b_ss = WB2B_0[WB2B_0["Episode"] == 0].copy()

        ss_merged = pd.merge(
            wbtv_ss,
            wb2b_ss[wb2b_pull_cols],
            on=["Season", "Episode"],
            how="left",
            suffixes=("", "_wb2b")
        )

        # -------------------------------------------------
        # 2️⃣ EPISODES — SEMANTIC
        # -------------------------------------------------
        ep_merged, matches_df, used_right, model, cross_model = (
            self.semantic_episode_match_and_merge(
                left_df=wbtv_ep,
                right_df=WB2B_0,
                normalize_title_fn=self.normalize_title,
                left_id_col="MPM Number",
                left_title_col="*Title name",
                right_title_col="wb2b epiosde name",
                right_pull_cols=wb2b_pull_cols,
                top_k=top_k,
                ce_threshold=ce_threshold,
                exclude_left_episode_zero=True,
                model=model,
                cross_model=cross_model,
            )
        )

        # -------------------------------------------------
        # 3️⃣ CONCAT + RESTORE ORDER ✅
        # -------------------------------------------------
        final = pd.concat([ss_merged, ep_merged], axis=0)
        final = final.sort_values("_wbtv_order").drop(columns="_wbtv_order").reset_index(drop=True)

        return final, matches_df, used_right, model, cross_model


    # ============================================================
    # ✅ MAIN ORCHESTRATOR (YOUR CORE REQUIREMENT)
    # ============================================================
    # ============================================================
    # ✅ MAIN ORCHESTRATOR (SERIES + MOVIE)
    # ============================================================
    def run_template_pipeline(self,
        input_folder: str = None,
        *,
        content_type: str = None,
        top_k: int = 5,
        ce_threshold: float = 0.75,
    ):
        """
        content_type:
        - series
        - movie
        """

        if content_type is None:
            content_type = self.ask_content_type()

        if input_folder is None:
            input_folder = self.pick_input_folder()
        
        if not input_folder:
            raise ValueError("No input folder selected.")


        sources = self.detect_sources(input_folder)

        Foundry = Series_season = Episodes = Movies = None
        if "FOUNDRY" in sources:
            Foundry, Series_season, Episodes, Movies = self.load_foundry(
                sources["FOUNDRY"], content_type=content_type
            )

        WBTV_0 = None
        WB2B_0 = None

        if "WBTV" in sources:
            WBTV_0 = self.load_wbtv(
                sources["WBTV"],
                content_type=content_type,
                series_season=Series_season if content_type == "series" else None
            )

        if "WB2B" in sources:
            WB2B_0 = self.load_wb2b(
                sources["WB2B"],
                content_type=content_type,
                series_season=Series_season if content_type == "series" else None,
                has_wbtv=("WBTV" in sources)
            )

        model = SentenceTransformer(MODEL_PATH)
        cross_model = CrossEncoder(CROSS_PATH)

        # =====================================================
        # SERIES ROUTING
        # =====================================================
        if content_type == "series":
            if ("WBTV" in sources) and ("WB2B" in sources) and ("FOUNDRY" in sources):
                master, _, _, model, cross_model = self.match_wbtv_with_wb2b_episodes(
                    WBTV_0, WB2B_0, model=model, cross_model=cross_model,
                    top_k=top_k, ce_threshold=ce_threshold
                )
                final_df, _, _, _, _ = self.match_wbtv_to_foundry_episodes(
                    master, Episodes, model=model, cross_model=cross_model,
                    top_k=top_k, ce_threshold=ce_threshold
                )
                return final_df

            if ("WBTV" in sources) and ("FOUNDRY" in sources) and ("WB2B" not in sources):
                final_df, _, _, _, _ = self.match_wbtv_to_foundry_episodes(
                    WBTV_0, Episodes, model=model, cross_model=cross_model,
                    top_k=top_k, ce_threshold=ce_threshold
                )
                return final_df

            if ("WB2B" in sources) and ("FOUNDRY" in sources) and ("WBTV" not in sources):
                final_df, _, _, _, _ = self.match_wb2b_to_foundry_episodes(
                    WB2B_0, Episodes, model=model, cross_model=cross_model,
                    top_k=top_k, ce_threshold=ce_threshold
                )
                return final_df

            if ("WBTV" in sources) and ("WB2B" in sources) and ("FOUNDRY" not in sources):
                final_df, _, _, _, _ = self.match_wbtv_with_wb2b_episodes(
                    WBTV_0, WB2B_0, model=model, cross_model=cross_model,
                    top_k=top_k, ce_threshold=ce_threshold
                )
                return final_df

            if ("WBTV" in sources) and (len(sources) == 1):
                return WBTV_0

            if ("WB2B" in sources) and (len(sources) == 1):
                return WB2B_0

            if ("FOUNDRY" in sources) and (len(sources) == 1):
                return Foundry.copy()

        # =====================================================
        # MOVIE ROUTING
        # =====================================================
        else:
            if ("WBTV" in sources) and ("WB2B" in sources) and ("FOUNDRY" in sources):
                master, _, _, model, cross_model = self.match_wbtv_with_wb2b_movies(
                    WBTV_0, WB2B_0, model=model, cross_model=cross_model,
                    top_k=top_k, ce_threshold=ce_threshold
                )
                final_df, _, _, _, _ = self.match_wbtv_to_foundry_movies(
                    master, Movies, model=model, cross_model=cross_model,
                    top_k=top_k, ce_threshold=ce_threshold
                )
                return final_df

            if ("WBTV" in sources) and ("FOUNDRY" in sources) and ("WB2B" not in sources):
                final_df, _, _, _, _ = self.match_wbtv_to_foundry_movies(
                    WBTV_0, Movies, model=model, cross_model=cross_model,
                    top_k=top_k, ce_threshold=ce_threshold
                )
                return final_df

            if ("WB2B" in sources) and ("FOUNDRY" in sources) and ("WBTV" not in sources):
                final_df, _, _, _, _ = self.match_wb2b_to_foundry_movies(
                    WB2B_0, Movies, model=model, cross_model=cross_model,
                    top_k=top_k, ce_threshold=ce_threshold
                )
                return final_df

            if ("WBTV" in sources) and ("WB2B" in sources) and ("FOUNDRY" not in sources):
                final_df, _, _, _, _ = self.match_wbtv_with_wb2b_movies(
                    WBTV_0, WB2B_0, model=model, cross_model=cross_model,
                    top_k=top_k, ce_threshold=ce_threshold
                )
                return final_df

            if ("WBTV" in sources) and (len(sources) == 1):
                return WBTV_0

            if ("WB2B" in sources) and (len(sources) == 1):
                return WB2B_0

            if ("FOUNDRY" in sources) and (len(sources) == 1):
                return Movies.copy()

        raise ValueError("No valid sources detected. Filenames must contain: WBTV / WB2B / Foundry")

    # =========================
    # PATH CONFIGURATION
    # =========================
    @staticmethod
    def pick_multilang_folder(title="Select folder with Multilanguages from Foundry"):
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)   # brings dialog to front (Windows helpful)
        root.update()                      # IMPORTANT: forces Tk to initialize properly
        folder = filedialog.askdirectory(title=title)
        root.destroy()
        return folder



    # =========================
    # FUNCTION: CHAR LIMIT PICKER
    # =========================
    @staticmethod
    def char_limit(target, df, en="no"):
        if en == "yes":
            preferred_cols = [
                "Synopsis (full)", "Synopsis (log line)", "Synopsis (180)",
                "Synopsis (60)", "Synopsis (60)_wb2b", "Synopsis (full)_wb2b",
                "description-400", "listings-description",
                "long-description", "main-description", "short-description"
            ]
        else:
            preferred_cols = [
                "description-400", "listings-description",
                "long-description", "main-description", "short-description"
            ]

        cols = [c for c in preferred_cols if c in df.columns]
        output = []

        for _, row in df.iterrows():
            # Collect valid non-null values
            values = {
                c: str(row[c]).strip()
                for c in cols
                if pd.notna(row[c]) and str(row[c]).strip()
            }

            # ✅ Strict filter: ONLY values within limit
            valid = [v for v in values.values() if 0 < len(v) <= target]

            if valid:
                # pick the longest within limit (best utilization)
                output.append(max(valid, key=len))
            else:
                # ❌ STRICT: no fallback allowed
                output.append("")

        return output

   
    @staticmethod
    def format_sheet_openpyxl(ws):
        """
        Apply header formatting, borders, auto width,
        and highlight invalid synopsis cells dynamically
        based on headers containing:
        (250 Character Limit), (500 Character Limit), etc.
        """

        import re
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        # =========================
        # STYLES
        # =========================
        header_fill = PatternFill(
            start_color="C6EFCE",
            end_color="C6EFCE",
            fill_type="solid"
        )

        header_font = Font(
            bold=True,
            color="000000"
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        normal_alignment = Alignment(
            vertical="top",
            wrap_text=True
        )

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        error_fill = PatternFill(
            start_color="FFC7CE",
            end_color="FFC7CE",
            fill_type="solid"
        )

        # =========================
        # HEADER FORMATTING
        # =========================
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        ws.freeze_panes = "A2"

        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

        # =========================
        # IDENTIFY DYNAMIC VALIDATION COLUMNS
        # =========================
        validation_columns = {}

        for cell in ws[1]:
            header = str(cell.value).strip() if cell.value else ""

            # Example:
            # Synopsis (250) SOURCE (250 Character Limit)
            # Synopsis (500) TRANSLATION (500 Character Limit)
            match = re.search(r"\((\d+)\s*Character\s*Limit\)\s*$", header)

            if match:
                limit = int(match.group(1))
                validation_columns[cell.column] = limit

        # =========================
        # APPLY BORDER + VALIDATION
        # =========================
        highlighted_count = 0

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = normal_alignment

                if cell.column in validation_columns:
                    limit = validation_columns[cell.column]
                    value = str(cell.value).strip() if cell.value is not None else ""

                    if not value or len(value) > limit:
                        cell.fill = error_fill
                        highlighted_count += 1

        # =========================
        # AUTO COLUMN WIDTH
        # =========================
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        print(
            f"Formatting applied on sheet '{ws.title}'. "
            f"Validation columns found: {len(validation_columns)}. "
            f"Highlighted cells: {highlighted_count}"
        )
    # =========================
    # HELPER FUNCTIONS
    # =========================
    @staticmethod
    def clean_text(series):
        return (
            series.astype(str)
            .str.replace(r'_x000D_', '', regex=False)
            .str.replace(r'_x000A_', '', regex=False)
            .str.strip()
        )
    
    @staticmethod
    def availability_status(series):
        if series.isna().all():
            return "Not available"

        non_empty = series.notna() & (series.astype(str).str.strip() != "")

        if non_empty.all():
            return "Fully available"
        elif non_empty.any():
            return "Partially available"
        else:
            return "Not available"

    def synopsis_status(self,series, max_len):
        if series.isna().all():
            return "Not available"

        cleaned = self.clean_text(series)
        lengths = cleaned.str.len()

        if (lengths > max_len).any():
            return "Exceed"
        elif (lengths < 5).any():
            return "Synopsis Missing"
        elif series.notna().all():
            return "All Synopsis Available"
        else:
            return "Synopsis Missing"
    
    @staticmethod
    def conditional_word_count(base_col, dependent_col):
        return sum(
            len(str(text).split())
            for text, dep in zip(base_col, dependent_col)
            if pd.isna(dep)
        )

