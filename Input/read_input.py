import openpyxl
import re

wb = openpyxl.load_workbook(r'C:\Users\mshanmugam\OneDrive - Warner Bros. Discovery\JUPYTER_PY\AI Projects\Smart Title Matching (AI-Based)\Input\Standalones.xlsx')
ws = wb.active
# Extract all Features values and parse AKA
features = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[1]:
        raw = str(row[1]).strip()
        # Parse: "Main Title (AKA. Alt Title)"
        aka_match = re.match(r'^(.+?)\s*\(AKA\.\s*(.+?)\s*\)$', raw)
        if aka_match:
            main = aka_match.group(1).strip()
            aka = aka_match.group(2).strip()
            features.append((raw, main, aka))
        else:
            features.append((raw, raw, None))

# Build SQL values for matching
# We'll match on CONVENIENCE_NAME against both main title and AKA
all_titles = set()
for raw, main, aka in features:
    all_titles.add(main)
    if aka:
        all_titles.add(aka)

# Print SQL-safe values
print(f"Total features: {len(features)}")
print(f"Unique search terms: {len(all_titles)}")
# Output as comma-separated for SQL IN clause
for t in sorted(all_titles):
    escaped = t.replace("'", "''")
    print(f"'{escaped}'")
